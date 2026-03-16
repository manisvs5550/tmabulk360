import os
import asyncio
import base64
import io
from datetime import date, datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, send_file, make_response, session
from playwright.async_api import async_playwright
from playwright.sync_api import sync_playwright
import pandas as pd
import sqlite3

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(32))

DB_PATH = "ship_movement.sqlite3"

VESSEL_NAMES = [
    "Abtenauer", "Abyssinian", "Amelie", "Andalucia", "Ardennes", "Asturcon", "Azteca",
    "Broompark", "Glenpark", "Jasmund", "Lefkada", "Mountpark", "Nordic Malmoe", "Sumatra"
]

# --- WARRANTED CONSUMPTION DICTS ---
WARRANTED_CONS_LADEN_13_5 = {
    "Abtenauer": 20.5, "Abyssinian": 21, "Amelie": 21.5, "Andalucia": 21.5, "Ardennes": 20.5, "Asturcon": 21, "Azteca": 21,
    "Broompark": 19.5, "Glenpark": 22, "Jasmund": 20.2, "Lefkada": 20, "Mountpark": 24, "Nordic Malmoe": 22, "Sumatra": 22.5
}
WARRANTED_CONS_LADEN_12 = {
    "Abtenauer": 15.75, "Abyssinian": 16.25, "Amelie": 16.5, "Andalucia": 16.5, "Ardennes": 15.75, "Asturcon": 16.25, "Azteca": 16.25,
    "Broompark": 14.5, "Glenpark": 16.9, "Jasmund": 15.6, "Lefkada": 15, "Mountpark": 16.9, "Nordic Malmoe": 16.9, "Sumatra": 17.2
}
WARRANTED_CONS_BALLAST_13_5 = {
    "Abtenauer": 18.5, "Abyssinian": 18.5, "Amelie": 19, "Andalucia": 19, "Ardennes": 18.5, "Asturcon": 18.5, "Azteca": 18.5,
    "Broompark": 18.5, "Glenpark": 20, "Jasmund": 18.8, "Lefkada": 18, "Mountpark": 21.5, "Nordic Malmoe": 20, "Sumatra": 21
}
WARRANTED_CONS_BALLAST_12 = {
    "Abtenauer": 14.25, "Abyssinian": 14.25, "Amelie": 14.5, "Andalucia": 14.5, "Ardennes": 14.25, "Asturcon": 14.25, "Azteca": 14.25,
    "Broompark": 13.3, "Glenpark": 14.4, "Jasmund": 14.7, "Lefkada": 15, "Mountpark": 14.4, "Nordic Malmoe": 14.4, "Sumatra": 15.2
}

# -- Warranted Speed (only exceptions) --
WARRANTED_SPEED_BALLAST_12 = {"Lefkada": 12.5, "Mountpark": 12}
WARRANTED_SPEED_BALLAST_13_5 = {"Mountpark": 13.5}
WARRANTED_SPEED_LADEN_12 = {}
WARRANTED_SPEED_LADEN_13_5 = {}

# Ensure uploads folder exists
if not os.path.exists("uploads"):
    os.makedirs("uploads")

# SQLite connection
def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# Import AIS Data
def import_ais_data(file_path, from_date, to_date):
    """Import AIS data from Excel file with proper date handling"""
    try:
        df = pd.read_excel(file_path)
        
        print("\n" + "=" * 60)
        print("IMPORT AIS DATA")
        print("=" * 60)
        print(f"Total rows in Excel: {len(df)}")
        print(f"Date range requested: {from_date} to {to_date}")
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower()
        column_map = {
            'date and time': 'Date and Time',
            'longitude': 'Longitude',
            'latitude': 'Latitude',
            'current direction': 'Current Direction',
            'current speed': 'Current Speed',
            'wave height': 'Wave Height',
            'wave period': 'Wave Period',
            'sig wave height': 'Sig Wave Height',
            'swell direction': 'Swell Direction',
            'swell height': 'Swell Height',
            'swell period': 'Swell Period',
            'wind direction 10m': 'Wind Direction 10m',
            'wind speed 10m': 'Wind Speed 10m',
            'gps distance': 'GPS Distance',
            'gps sog': 'GPS SOG',
            'heading': 'Heading',
            'current factor': 'Current Factor',
            'weather factor': 'Weather Factor',
            'dt performance speed': 'DT Performance Speed'
        }
        
        existing_cols = [col for col in column_map.keys() if col in df.columns]
        print(f"✓ Matched {len(existing_cols)} columns")
        
        if not existing_cols:
            return "❌ No matching columns found in AIS file"
        
        df = df[existing_cols]
        df = df.rename(columns=column_map)
        
        if 'Current Direction' in df.columns:
            df = df[df['Current Direction'] > 0]
            print(f"✓ Current Direction > 0 filter: {len(df)} rows")
        
        if 'Date and Time' not in df.columns:
            return "❌ 'Date and Time' column not found"
        
        df['Date and Time'] = pd.to_datetime(df['Date and Time'], errors='coerce')
        
        before_drop = len(df)
        df = df.dropna(subset=['Date and Time'])
        if before_drop != len(df):
            print(f"⚠️ Removed {before_drop - len(df)} invalid dates")
        
        if len(df) == 0:
            return "❌ All dates invalid in AIS file"
        
        print(f"Date range in file: {df['Date and Time'].min()} to {df['Date and Time'].max()}")
        
        from_date_parsed = pd.to_datetime(from_date)
        to_date_parsed = pd.to_datetime(to_date)
        
        mask = (df['Date and Time'] >= from_date_parsed) & (df['Date and Time'] <= to_date_parsed)
        df = df[mask]
        print(f"After date filter: {len(df)} rows")
        
        if df.empty:
            return f"⚠️ No AIS data in date range {from_date_parsed.date()} to {to_date_parsed.date()}"
        
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute("DROP TABLE IF EXISTS AISData")
            cursor.execute("""
                CREATE TABLE AISData (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    `Date and Time` TEXT,
                    `Longitude` REAL,
                    `Latitude` REAL,
                    `Current Direction` REAL,
                    `Current Speed` REAL,
                    `Wave Height` REAL,
                    `Wave Period` REAL,
                    `Sig Wave Height` REAL,
                    `Swell Direction` REAL,
                    `Swell Height` REAL,
                    `Swell Period` REAL,
                    `Wind Direction 10m` REAL,
                    `Wind Speed 10m` REAL,
                    `GPS Distance` REAL,
                    `GPS SOG` REAL,
                    `Heading` REAL,
                    `Current Factor` REAL,
                    `Weather Factor` REAL,
                    `DT Performance Speed` REAL
                )
            """)
            
            for _, row in df.iterrows():
                cursor.execute("""
                    INSERT INTO AISData
                    (`Date and Time`, `Longitude`, `Latitude`, `Current Direction`, `Current Speed`,
                     `Wave Height`, `Wave Period`, `Sig Wave Height`, `Swell Direction`, `Swell Height`,
                     `Swell Period`, `Wind Direction 10m`, `Wind Speed 10m`, `GPS Distance`, `GPS SOG`,
                     `Heading`, `Current Factor`, `Weather Factor`, `DT Performance Speed`)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    row.get('Date and Time').isoformat() if pd.notnull(row.get('Date and Time')) else None,
                    row.get('Longitude'),
                    row.get('Latitude'),
                    row.get('Current Direction'),
                    row.get('Current Speed'),
                    row.get('Wave Height'),
                    row.get('Wave Period'),
                    row.get('Sig Wave Height'),
                    row.get('Swell Direction'),
                    row.get('Swell Height'),
                    row.get('Swell Period'),
                    row.get('Wind Direction 10m'),
                    row.get('Wind Speed 10m'),
                    row.get('GPS Distance'),
                    row.get('GPS SOG'),
                    row.get('Heading'),
                    row.get('Current Factor'),
                    row.get('Weather Factor'),
                    row.get('DT Performance Speed')
                ))
            
            conn.commit()
        
        print(f"✅ SUCCESS: {len(df)} AIS rows imported")
        return f"✅ AIS data uploaded successfully ({len(df)} rows)"
        
    except Exception as e:
        import traceback
        print("❌ ERROR:")
        print(traceback.format_exc())
        return f"❌ Error importing AIS data: {e}"

# Import VesonData
def import_weather_data(file_path, from_date, to_date):
    """Import Veson data - accepts any vessel name from Excel"""
    try:
        df = pd.read_excel(file_path)
        
        print("\n" + "=" * 60)
        print("IMPORT WEATHER DATA")
        print("=" * 60)
        print(f"Total rows in Excel: {len(df)}")
        print(f"Date range requested: {from_date} to {to_date}")
        
        df.columns = df.columns.str.strip().str.lower()
        column_map = {
            'date time': 'Date Time',
            'steaming hours': 'Steaming Hours',
            'observed distance': 'Observed Distance',
            'cp ordered speed': 'CP Ordered Speed',
            'projected speed': 'Projected Speed',
            'average rpm': 'Average RPM',
            'lsf consumption': 'LSF Consumption',
            'lgo consumption': 'LGO Consumption',
            'vessel condition': 'Vessel Condition',
            'sea height': 'Sea Height',
            'vessel name': 'Vessel Name',
            'main engine hrs': 'Main Engine Hrs'
        }
        
        existing_cols = [col for col in column_map.keys() if col in df.columns]
        print(f"✓ Matched {len(existing_cols)} columns")
        
        df = df[existing_cols]
        df = df.rename(columns=column_map)
        
        if 'LSF Consumption' in df.columns and 'LGO Consumption' in df.columns:
            df['Total Consumption'] = df['LSF Consumption'].fillna(0) + df['LGO Consumption'].fillna(0)
            print("✅ Created 'Total Consumption'")
        elif 'Total Consumption' not in df.columns:
            return "❌ Cannot calculate Total Consumption"
        
        if 'Date Time' not in df.columns:
            return "❌ 'Date Time' column not found"
        
        df['Date Time'] = pd.to_datetime(df['Date Time'], errors='coerce')
        
        before_drop = len(df)
        df = df.dropna(subset=['Date Time'])
        if before_drop != len(df):
            print(f"⚠️ Removed {before_drop - len(df)} invalid dates")
        
        if len(df) == 0:
            return "❌ All dates invalid"
        
        print(f"Date range in file: {df['Date Time'].min()} to {df['Date Time'].max()}")
        
        from_date_parsed = pd.to_datetime(from_date)
        to_date_parsed = pd.to_datetime(to_date)
        
        mask = (df['Date Time'] >= from_date_parsed) & (df['Date Time'] <= to_date_parsed)
        df = df[mask]
        print(f"After date filter: {len(df)} rows")
        
        if len(df) == 0:
            return f"⚠️ No data in range {from_date_parsed.date()} to {to_date_parsed.date()}"
        
        # Clean vessel names
        if 'Vessel Name' in df.columns:
            df['Vessel Name'] = df['Vessel Name'].str.strip().str.title()
            print(f"Vessels in file: {df['Vessel Name'].unique()}")
            df = df[df['Vessel Name'].notna() & (df['Vessel Name'] != '')]
            print(f"After removing empty vessel names: {len(df)} rows")
        
        if df.empty:
            return "⚠️ No valid data in date range"
        
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute("DROP TABLE IF EXISTS VesonData")
            cursor.execute("""
                CREATE TABLE VesonData (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    `Date Time` TEXT,
                    `Steaming Hours` REAL,
                    `Observed Distance` REAL,
                    `CP Ordered Speed` REAL,
                    `Projected Speed` REAL,
                    `Average RPM` REAL,
                    `LSF Consumption` REAL,
                    `LGO Consumption` REAL,
                    `Vessel Condition` TEXT,
                    `Sea Height` REAL,
                    `Total Consumption` REAL,
                    `Vessel Name` TEXT,
                    `Main Engine Hrs` REAL,
                    `hind_cast_avg_cf` REAL,
                    `Current adjusted +-distance` REAL,
                    `Hindcast-Max SWH` REAL,
                    `Hindcast-Max Wind Speed` REAL,
                    `Hind Cast-Good weather` TEXT
                )
            """)
            
            for _, row in df.iterrows():
                cursor.execute("""
                    INSERT INTO VesonData
                    (`Date Time`, `Steaming Hours`, `Observed Distance`, `CP Ordered Speed`, `Projected Speed`,
                     `Average RPM`, `LSF Consumption`, `LGO Consumption`, `Vessel Condition`, `Sea Height`,
                     `Total Consumption`, `Vessel Name`, `Main Engine Hrs`, `hind_cast_avg_cf`, 
                     `Current adjusted +-distance`, `Hindcast-Max SWH`, `Hindcast-Max Wind Speed`, `Hind Cast-Good weather`)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    row.get('Date Time').isoformat() if pd.notnull(row.get('Date Time')) else None,
                    row.get('Steaming Hours'),
                    row.get('Observed Distance'),
                    row.get('CP Ordered Speed'),
                    row.get('Projected Speed'),
                    row.get('Average RPM'),
                    row.get('LSF Consumption'),
                    row.get('LGO Consumption'),
                    row.get('Vessel Condition'),
                    row.get('Sea Height'),
                    row.get('Total Consumption'),
                    row.get('Vessel Name'),
                    row.get('Main Engine Hrs'),
                    None, None, None, None, None
                ))
            
            conn.commit()
        
        vessel_name = df['Vessel Name'].iloc[0] if 'Vessel Name' in df.columns else "Unknown"
        print(f"✅ SUCCESS: {len(df)} rows for {vessel_name}")
        
        return f"✅ Veson data uploaded for {vessel_name} ({len(df)} rows)"
        
    except Exception as e:
        import traceback
        print("❌ ERROR:")
        print(traceback.format_exc())
        return f"❌ Error importing Veson data: {e}"

# Calculate hindcast values
def calculate_hindcast_and_adjusted_distance():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT id, `Date Time`, `Steaming Hours` FROM VesonData")
        veson_rows = cursor.fetchall()

        for row in veson_rows:
            veson_id, date_time_str, steaming_hours = row
            avg_cf, adjusted_distance = None, None
            hindcast_max_swh, hindcast_max_wind, good_weather = None, None, None

            if date_time_str and steaming_hours:
                date_time = pd.to_datetime(date_time_str)
                start_time = date_time - timedelta(hours=float(steaming_hours))
                end_time = date_time

                # Hind cast avg CF
                cursor.execute("""
                    SELECT AVG(`Current Factor`) FROM AISData
                    WHERE `Date and Time` BETWEEN ? AND ?
                """, (start_time.isoformat(), end_time.isoformat()))
                avg_cf = cursor.fetchone()[0]
                if avg_cf is not None:
                    avg_cf = round(avg_cf, 2)
                    adjusted_distance = round(avg_cf * float(steaming_hours), 2)

                # Hindcast-Max SWH
                cursor.execute("""
                    SELECT MAX(`Sig Wave Height`) FROM AISData
                    WHERE `Date and Time` BETWEEN ? AND ?
                """, (start_time.isoformat(), end_time.isoformat()))
                hindcast_max_swh = cursor.fetchone()[0]
                if hindcast_max_swh is not None:
                    hindcast_max_swh = round(hindcast_max_swh, 2)

                # Hindcast-Max Wind Speed
                cursor.execute("""
                    SELECT MAX(`Wind Speed 10m`) FROM AISData
                    WHERE `Date and Time` BETWEEN ? AND ?
                """, (start_time.isoformat(), end_time.isoformat()))
                hindcast_max_wind = cursor.fetchone()[0]
                if hindcast_max_wind is not None:
                    hindcast_max_wind = round(hindcast_max_wind, 2)

                # Good weather
                if hindcast_max_swh is not None and hindcast_max_wind is not None:
                    good_weather = "Yes" if (hindcast_max_swh < 1.501 and hindcast_max_wind < 16) else "No"

            cursor.execute("""
                UPDATE VesonData
                SET hind_cast_avg_cf = ?,
                    `Current adjusted +-distance` = ?,
                    `Hindcast-Max SWH` = ?,
                    `Hindcast-Max Wind Speed` = ?,
                    `Hind Cast-Good weather` = ?
                WHERE id = ?
            """, (avg_cf, adjusted_distance, hindcast_max_swh, hindcast_max_wind, good_weather, veson_id))

        conn.commit()
        conn.close()
        print("✅ Calculations completed.")

    except Exception as e:
        print(f"❌ Error calculating values: {e}")


@app.route("/ship-performance", methods=["GET", "POST"])
def ship_performance():
    if 'logged_in' not in session:
        return redirect(url_for('login'))

    ais_message = None
    weather_message = None
    vessel_selected = None

    if request.method == "POST":
        vessel_selected = session.get("vessel_name", None)
        form_type = request.form.get("form_type", "")
        
        if form_type == "upload":
            from_date = request.form.get("from_date")
            to_date = request.form.get("to_date")
            ais_file = request.files.get("ais_file")
            veson_file = request.files.get("weather_file")

            if not from_date or not to_date:
                weather_message = "❌ Please select both From Date and To Date"
                return render_template(
                    "upload.html",
                    ais_message=ais_message,
                    weather_message=weather_message,
                    selected_vessel=vessel_selected
                )

            ais_vessel = None
            veson_vessel = None

            # Upload AIS and extract vessel name
            if ais_file and ais_file.filename:
                path = os.path.join("uploads", ais_file.filename)
                ais_file.save(path)
                
                try:
                    df_ais = pd.read_excel(path)
                    df_ais.columns = df_ais.columns.str.strip().str.lower()
                    if "vessel name" in df_ais.columns:
                        unique_vessels = df_ais["vessel name"].dropna().unique()
                        
                        if len(unique_vessels) == 1:
                            ais_vessel = str(unique_vessels[0]).strip().title()
                        elif len(unique_vessels) > 1:
                            ais_message = f"❌ AIS file contains multiple vessels: {', '.join([str(v) for v in unique_vessels])}. Please upload data for only one vessel."
                            return render_template(
                                "upload.html",
                                ais_message=ais_message,
                                weather_message=weather_message,
                                selected_vessel=vessel_selected
                            )
                except Exception as e:
                    print(f"⚠️ Could not detect vessel from AIS file: {e}")
                
                if not ais_message or "✅" in ais_message:
                    ais_message = import_ais_data(path, from_date, to_date)
            
            # Upload Veson and extract vessel name
            if veson_file and veson_file.filename:
                path = os.path.join("uploads", veson_file.filename)
                veson_file.save(path)
                
                try:
                    df_veson = pd.read_excel(path)
                    df_veson.columns = df_veson.columns.str.strip().str.lower()
                    if "vessel name" in df_veson.columns:
                        unique_vessels = df_veson["vessel name"].dropna().unique()
                        
                        if len(unique_vessels) == 1:
                            veson_vessel = str(unique_vessels[0]).strip().title()
                        elif len(unique_vessels) > 1:
                            weather_message = f"❌ Veson file contains multiple vessels: {', '.join([str(v) for v in unique_vessels])}. Please upload data for only one vessel."
                            return render_template(
                                "upload.html",
                                ais_message=ais_message,
                                weather_message=weather_message,
                                selected_vessel=vessel_selected
                            )
                except Exception as e:
                    print(f"⚠️ Could not detect vessel from Veson file: {e}")
                
                # Validate vessel names match between files
                if ais_vessel and veson_vessel and ais_vessel != veson_vessel:
                    weather_message = f"❌ Vessel name mismatch! AIS file has '{ais_vessel}' but Veson file has '{veson_vessel}'. Please upload files for the same vessel."
                    return render_template(
                        "upload.html",
                        ais_message=ais_message,
                        weather_message=weather_message,
                        selected_vessel=vessel_selected
                    )
                
                # Store vessel name in session
                if veson_vessel:
                    vessel_selected = veson_vessel
                    session["vessel_name"] = vessel_selected
                elif ais_vessel:
                    vessel_selected = ais_vessel
                    session["vessel_name"] = vessel_selected
                
                if not weather_message or "✅" in weather_message:
                    weather_message = import_weather_data(path, from_date, to_date)
                    
                    if weather_message and "✅" in weather_message:
                        calculate_hindcast_and_adjusted_distance()
                        
                        if not vessel_selected:
                            try:
                                conn = get_db_connection()
                                cursor = conn.cursor()
                                cursor.execute("SELECT DISTINCT `Vessel Name` FROM VesonData LIMIT 1")
                                result = cursor.fetchone()
                                if result:
                                    vessel_selected = result[0]
                                    session["vessel_name"] = vessel_selected
                                conn.close()
                            except:
                                pass

    return render_template(
        "upload.html",
        ais_message=ais_message,
        weather_message=weather_message,
        selected_vessel=vessel_selected
    )

# Helper: build summary data from DB
def _build_summary(vessel_name):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DISTINCT `Vessel Condition` FROM VesonData
        WHERE `Vessel Name` = ?
          AND `Vessel Condition` IS NOT NULL 
          AND `Vessel Condition` != ''
    """, [vessel_name])
    vessel_conditions = [row[0] for row in cursor.fetchall()]
    summary = {}

    for vc in vessel_conditions:
        summary[vc] = {}
        for speed in [13.5, 12.0]:
            for weather_type, weather_filter in [("all_weather", None), ("good_weather", "Yes")]:
                key = f"{speed}_{weather_type}"

                conditions = [
                    "`Vessel Name` = ?",
                    "`Vessel Condition` = ?",
                    "`CP Ordered Speed` = ?",
                    "`Steaming Hours` > 20",
                    "`Main Engine Hrs` > 20"
                ]
                params = [vessel_name, vc, speed]

                if weather_filter:
                    conditions.append("`Hind Cast-Good weather` = ?")
                    params.append(weather_filter)

                where_clause = " AND ".join(conditions)

                cursor.execute(f"SELECT SUM(`Steaming Hours`) FROM VesonData WHERE {where_clause}", params)
                total_time = cursor.fetchone()[0] or 0

                cursor.execute(f"SELECT SUM(`Observed Distance`) FROM VesonData WHERE {where_clause}", params)
                total_distance = cursor.fetchone()[0] or 0

                cursor.execute(f"SELECT SUM(`Total Consumption`) FROM VesonData WHERE {where_clause}", params)
                total_consumption = cursor.fetchone()[0] or 0

                avg_speed = round(total_distance / total_time, 2) if total_time > 0 else 0
                avg_consumption_per_day = round((total_consumption / total_time) * 24, 2) if total_time > 0 else 0

                if weather_type == "good_weather":
                    cursor.execute(
                        f"""SELECT SUM(`Steaming Hours` * IFNULL(`hind_cast_avg_cf`, 0))
                            FROM VesonData WHERE {where_clause}
                        """, params
                    )
                    total_applicable_current_factor = cursor.fetchone()[0] or 0
                else:
                    total_applicable_current_factor = None

                summary[vc][key] = {
                    "total_time": round(total_time, 2),
                    "total_distance": round(total_distance, 2),
                    "total_consumption": round(total_consumption, 2),
                    "avg_speed": avg_speed,
                    "avg_consumption_per_day": avg_consumption_per_day,
                    "total_applicable_current_factor": round(total_applicable_current_factor, 2) if total_applicable_current_factor is not None else None
                }

    conn.close()
    return summary


# Generate performance summary
@app.route("/generate_summary", methods=["POST"])
def generate_summary():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    vessel_name = request.form.get("vessel_name")
    
    if not vessel_name:
        return render_template(
            "upload.html",
            ais_message=None,
            weather_message=None,
            selected_vessel=None,
            error="No vessel data found. Please upload data first.",
        )
    
    session["vessel_name"] = vessel_name
    
    try:
        summary = _build_summary(vessel_name)
        
        x_labels = ['12 Ballast', '13.5 Ballast', '12 Laden', '13.5 Laden']
        
        warranted_speeds = [
            WARRANTED_SPEED_BALLAST_12.get(vessel_name, 12),
            WARRANTED_SPEED_BALLAST_13_5.get(vessel_name, 13.5),
            WARRANTED_SPEED_LADEN_12.get(vessel_name, 12),
            WARRANTED_SPEED_LADEN_13_5.get(vessel_name, 13.5)
        ]
        warranted_cons_values = [
            WARRANTED_CONS_BALLAST_12.get(vessel_name, 0),
            WARRANTED_CONS_BALLAST_13_5.get(vessel_name, 0),
            WARRANTED_CONS_LADEN_12.get(vessel_name, 0),
            WARRANTED_CONS_LADEN_13_5.get(vessel_name, 0)
        ]
        
        observed_speeds = [
            summary.get('Ballast', {}).get('12.0_good_weather', {}).get('avg_speed', 0),
            summary.get('Ballast', {}).get('13.5_good_weather', {}).get('avg_speed', 0),
            summary.get('Laden', {}).get('12.0_good_weather', {}).get('avg_speed', 0),
            summary.get('Laden', {}).get('13.5_good_weather', {}).get('avg_speed', 0)
        ]
        observed_cons_values = [
            summary.get('Ballast', {}).get('12.0_good_weather', {}).get('avg_consumption_per_day', 0),
            summary.get('Ballast', {}).get('13.5_good_weather', {}).get('avg_consumption_per_day', 0),
            summary.get('Laden', {}).get('12.0_good_weather', {}).get('avg_consumption_per_day', 0),
            summary.get('Laden', {}).get('13.5_good_weather', {}).get('avg_consumption_per_day', 0)
        ]
        percent_changes = [
            round((o - w), 2) if (o and w) else 0
            for o, w in zip(observed_speeds, warranted_speeds)
        ]
        
        return render_template(
            "summary.html",
            summary=summary,
            x_labels=x_labels,
            warranted_speeds=warranted_speeds,
            warranted_cons_values=warranted_cons_values,
            observed_speeds=observed_speeds,
            observed_cons_values=observed_cons_values,
            percent_changes=percent_changes,
            vessel_name=vessel_name
        )
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return render_template(
            "upload.html",
            ais_message=None,
            weather_message=None,
            selected_vessel=vessel_name,
            error=f"Error generating summary: {e}",
        )


# Download summary as PDF
def _get_logo_b64():
    """Read logo.png and return base64-encoded string."""
    logo_path = os.path.join(app.static_folder, "images", "logo.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    return ""


@app.route("/download_summary_pdf")
def download_summary_pdf():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    try:
        vessel_name = session.get("vessel_name", VESSEL_NAMES[0])
        summary = _build_summary(vessel_name)
        html = render_template(
            "summary_pdf.html",
            summary=summary,
            logo_b64=_get_logo_b64(),
            generated_date=date.today().strftime("%d %b %Y"),
            vessel_name=vessel_name,
        )

        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.set_content(html, wait_until="networkidle")
            pdf_bytes = page.pdf(
                format="A4",
                print_background=True,
                margin={"top": "1cm", "bottom": "1cm", "left": "1cm", "right": "1cm"}
            )
            browser.close()

        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        safe_name = vessel_name.replace(' ', '_')
        response.headers['Content-Disposition'] = f'attachment; filename=Performance_Summary_{safe_name}.pdf'
        return response
    except Exception as e:
        return f"❌ Error generating PDF: {e}"


@app.route("/generate_ship_report", methods=["GET"])
def generate_ship_report():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        vessel_name = session.get("vessel_name", VESSEL_NAMES[0])
        
        conn = sqlite3.connect(DB_PATH)
        veson = pd.read_sql_query("SELECT * FROM VesonData WHERE `Vessel Name` = ?", conn, params=[vessel_name])
        ais = pd.read_sql_query("SELECT * FROM AISData", conn)
        conn.close()
        
        # Filter by Steaming Hours and Main Engine Hrs
        veson = veson[(veson["Steaming Hours"] > 20) & (veson["Main Engine Hrs"] > 20)]
        
        def wind_speed_to_beaufort(wind_knots):
            """Convert wind speed in knots to Beaufort Force (0-12)"""
            if wind_knots is None or pd.isna(wind_knots):
                return None
            if wind_knots < 1:
                return 0
            elif wind_knots < 4:
                return 1
            elif wind_knots < 7:
                return 2
            elif wind_knots < 11:
                return 3
            elif wind_knots < 17:
                return 4
            elif wind_knots < 22:
                return 5
            elif wind_knots < 28:
                return 6
            elif wind_knots < 34:
                return 7
            elif wind_knots < 41:
                return 8
            elif wind_knots < 48:
                return 9
            elif wind_knots < 56:
                return 10
            elif wind_knots < 64:
                return 11
            else:
                return 12
        
        output = []
        rows_with_actual_dt = []
        
        for ix, row in veson.iterrows():
            dt = pd.to_datetime(row["Date Time"])
            sh = row["Steaming Hours"]
            
            if pd.notnull(dt):
                excel_date = dt.strftime("%d/%m/%Y %H:%M")
            else:
                excel_date = ""
            
            if pd.isnull(dt) or pd.isnull(sh) or sh <= 0:
                avg_cf = None
                adj_dist = None
                max_swh = None
                max_wind = None
                good_weather = None
                avg_current_dir = None
                avg_wind_force = None
            else:
                start_time = dt - timedelta(hours=float(sh))
                end_time = dt
                mask = (pd.to_datetime(ais["Date and Time"]) >= start_time) & (pd.to_datetime(ais["Date and Time"]) < end_time)
                ais_window = ais[mask]
                
                avg_cf = ais_window["Current Factor"].mean() if not ais_window.empty and "Current Factor" in ais_window.columns else None
                adj_dist = avg_cf * sh if (avg_cf is not None) else None
                max_swh = ais_window["Sig Wave Height"].max() if not ais_window.empty and "Sig Wave Height" in ais_window.columns else None
                max_wind = ais_window["Wind Speed 10m"].max() if not ais_window.empty and "Wind Speed 10m" in ais_window.columns else None
                
                avg_current_dir = ais_window["Current Direction"].mean() if not ais_window.empty and "Current Direction" in ais_window.columns else None
                
                avg_wind_speed = ais_window["Wind Speed 10m"].mean() if not ais_window.empty and "Wind Speed 10m" in ais_window.columns else None
                avg_wind_force = wind_speed_to_beaufort(avg_wind_speed) if avg_wind_speed is not None else None
                
                if (max_swh is not None) and (max_wind is not None):
                    good_weather = "Yes" if (max_swh < 1.501 and max_wind < 16) else "No"
                else:
                    good_weather = None
            
            xrow = [
                excel_date,
                row.get("Steaming Hours"),
                row.get("Observed Distance"),
                row.get("CP Ordered Speed"),
                row.get("Projected Speed"),
                row.get("Total Consumption"),
                row.get("Vessel Condition"),
                row.get("Sea Height"),
                round(avg_current_dir, 1) if avg_current_dir is not None else "",
                avg_wind_force if avg_wind_force is not None else "",
                round(avg_cf, 2) if avg_cf is not None else "",
                round(adj_dist, 2) if adj_dist is not None else "",
                round(max_swh, 2) if max_swh is not None else "",
                round(max_wind, 2) if max_wind is not None else "",
                good_weather if good_weather is not None else "",
            ]
            rows_with_actual_dt.append((dt if pd.notnull(dt) else pd.Timestamp.min, xrow))
        
        rows_with_actual_dt.sort(key=lambda tup: tup[0])
        output = [tup[1] for tup in rows_with_actual_dt]
        
        headers = [
            "Date/Time", "Steaming Hours", "Obs Dist (NM)", "Ordered Speed (kts)", "Obs. Speed (kts)",
            "Total Con (Mt)", "Vessel Cond", "Sea Wave Height (m)", "Cur Direction (°)", "Wind (Bf)",
            "Avg CF (kn)", "CF adj dist-Nm (CF*Steaming hours)", "Max SWH (m)", "Max wind speed (knots)", "Good weather Yes/No"
        ]
        df_report = pd.DataFrame(output, columns=headers)
        
        output_stream = io.BytesIO()
        with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
            df_report.to_excel(writer, index=False, sheet_name='Ship Report')
        output_stream.seek(0)
        
        wb = openpyxl.load_workbook(output_stream)
        
        if "Executive" in wb.sheetnames:
            del wb["Executive"]
        ws_ex = wb.create_sheet("Executive", 0)
        
        ws_ex["A1"] = f"Vessel: {vessel_name}"
        ws_ex.merge_cells("A1:G1")
        ws_ex["A1"].font = Font(size=14, bold=True)
        ws_ex["A1"].alignment = Alignment(horizontal="left", vertical="center")
        
        headers_exec = [
            "Condition", "Warranted Speed (knots)", "Warranted Consumption (Mts)", "Observed speed (Knots)",
            "Observed Consumption (Mts)", "% Change in Observed Speed", "% Change in Observed Consumption"
        ]
        dark_blue = PatternFill(start_color="193354", end_color="193354", fill_type="solid")
        white_bold = Font(color="FFFFFF", bold=True)
        
        for col, value in enumerate(headers_exec, start=1):
            cell = ws_ex.cell(row=3, column=col, value=value)
            cell.fill = dark_blue
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_ex.row_dimensions[3].height = 22
        
        conn = get_db_connection()
        cursor = conn.cursor()
        summary_inputs = [
            ("Ballast", 12, WARRANTED_CONS_BALLAST_12, WARRANTED_SPEED_BALLAST_12, "12.0_good_weather"),
            ("Ballast", 13.5, WARRANTED_CONS_BALLAST_13_5, WARRANTED_SPEED_BALLAST_13_5, "13.5_good_weather"),
            ("Laden", 12, WARRANTED_CONS_LADEN_12, WARRANTED_SPEED_LADEN_12, "12.0_good_weather"),
            ("Laden", 13.5, WARRANTED_CONS_LADEN_13_5, WARRANTED_SPEED_LADEN_13_5, "13.5_good_weather"),
        ]
        build_summary = {}
        
        for cond in ["Ballast", "Laden"]:
            build_summary[cond] = {}
            for speed in [13.5, 12.0]:
                cursor.execute("""
                    SELECT 
                        SUM(`Observed Distance`) as total_distance,
                        SUM(`Steaming Hours`) as total_time,
                        SUM(`Total Consumption`) as total_consumption
                    FROM VesonData
                    WHERE `Vessel Name` = ?
                    AND `Vessel Condition` = ?
                    AND `CP Ordered Speed` = ?
                    AND `Steaming Hours` > 20
                    AND `Main Engine Hrs` > 20
                    AND `Vessel Condition` IS NOT NULL 
                    AND `Vessel Condition` != ''
                    AND `Hind Cast-Good weather` = 'Yes'
                """, [vessel_name, cond, speed])
                
                res = cursor.fetchone()
                key = f"{speed}_good_weather"
                
                if res and res[0] is not None and res[1] is not None and res[1] > 0:
                    avg_speed = round(res[0] / res[1], 2)
                    avg_cons = round((res[2] / res[1]) * 24, 2) if res[2] else 0
                else:
                    avg_speed = ""
                    avg_cons = ""
                
                build_summary[cond][key] = {
                    "avg_speed": avg_speed,
                    "avg_cons": avg_cons
                }
        
        conn.close()
        
        blue_fill = PatternFill(start_color="B7DCF6", end_color="B7DCF6", fill_type="solid")
        side = Side(border_style="thin", color="000000")
        
        for i, (cond, speed, warranted_cons_dict, warranted_speed_dict, summary_key) in enumerate(summary_inputs, start=4):
            warranted_speed = warranted_speed_dict.get(vessel_name, speed)
            warranted_cons = warranted_cons_dict.get(vessel_name, "")
            avg_speed = build_summary[cond][summary_key]["avg_speed"]
            avg_cons = build_summary[cond][summary_key]["avg_cons"]
            
            ws_ex.cell(row=i, column=1, value=cond)
            ws_ex.cell(row=i, column=2, value=warranted_speed)
            ws_ex.cell(row=i, column=3, value=warranted_cons)
            ws_ex.cell(row=i, column=4, value=avg_speed)
            ws_ex.cell(row=i, column=5, value=avg_cons)
            ws_ex.cell(row=i, column=6, value=f'=IF(B{i}=0,"",D{i}/B{i})')
            ws_ex.cell(row=i, column=6).number_format = '0.00%'
            ws_ex.cell(row=i, column=7, value=f'=IF(C{i}=0,"",E{i}/C{i})')
            ws_ex.cell(row=i, column=7).number_format = '0.00%'
            
            for col in range(1, 8):
                c = ws_ex.cell(row=i, column=col)
                if col in [2, 3, 4, 5]:
                    c.fill = blue_fill
                c.border = Border(left=side, right=side, top=side, bottom=side)
                c.alignment = Alignment(horizontal="center", vertical="center")
        
        ws = wb['Ship Report']
        green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type="solid")
        
        for row_idx in range(2, ws.max_row + 1):
            good_weather_cell = ws.cell(row=row_idx, column=15)
            if good_weather_cell.value == "Yes":
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = green_fill
        
        styled_stream = io.BytesIO()
        wb.save(styled_stream)
        styled_stream.seek(0)
        
        return send_file(
            styled_stream,
            as_attachment=True,
            download_name=f"Ship_Report_{vessel_name}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return f"❌ Error generating Ship Report: {e}"


@app.route("/generate-pdf-image")
async def generate_pdf_image():
    """Generates a PDF of the index page using a headless browser."""
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page()
        
        # Navigate to the live index page
        # Note: This requires the Flask app to be running
        base_url = request.url_root
        await page.goto(base_url, wait_until="networkidle")
        
        # Generate the PDF from the live page
        pdf_bytes = await page.pdf(format="A4", print_background=True, margin={"top": "1cm", "bottom": "1cm", "left": "1cm", "right": "1cm"})
        
        await browser.close()

    # Create a response
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=tmaops360_preview.pdf'
    return response


@app.route("/sw.js")
def service_worker():
    """Serve SW from root scope so it can cache all routes."""
    return send_from_directory(app.static_folder, "sw.js", mimetype="application/javascript")


@app.route("/offline")
def offline():
    return render_template("offline.html")


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form.get("username") == "Admin" and request.form.get("password") == "Admin":
            session['logged_in'] = True
            return redirect(url_for("dashboard"))
        else:
            return render_template("login.html", error="Invalid credentials")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))


@app.route("/dashboard")
def dashboard():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template("dashboard.html")


@app.route("/contact", methods=["POST"])
def contact():
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip()
    company = request.form.get("company", "").strip()
    reason = request.form.get("reason", "").strip()
    # TODO: Process the contact form (e.g., send email, save to DB)
    return redirect(url_for("index", _anchor="contact"))


if __name__ == "__main__":
    env = os.environ.get("FLASK_ENV", "development")
    if env == "production":
        from waitress import serve
        port = int(os.environ.get("PORT", 8080))
        print(f"Serving on http://0.0.0.0:{port}")
        serve(app, host="0.0.0.0", port=port)
    else:
        from hypercorn.config import Config
        from hypercorn.asyncio import serve
        from asgiref.wsgi import WsgiToAsgi

        config = Config()
        config.bind = ["localhost:5000"]
        config.use_reloader = True
        
        asyncio.run(serve(WsgiToAsgi(app), config))
